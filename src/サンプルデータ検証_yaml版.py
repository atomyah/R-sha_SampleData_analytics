#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
サンプルデータ検証_yaml版.py  （rates.yaml 読み込み版）

HRTech SaaS（matchinggood想定）からエクスポートした「R社サンプルデータ」形式の
Excelを読み込み、契約データ＋勤怠実績データから請求金額・支払金額を再計算し、
「サンプルデータ概要」シートの目標値（MG税込請求金額／MG税込支払金額）と突き合わせて
OK / NG を自動判定する。

★料率・源泉所得税の月額表・各種設定は、すべて同じフォルダの rates.yaml から読み込む。
  年度や都道府県が変わったら rates.yaml だけ直せばよく、このコードは触らなくてよい。

使い方:
    python サンプルデータ検証.py <Excelファイル>
    python サンプルデータ検証.py <Excelファイル> --rates rates.yaml --tax-year 2026

判定ルール:
  ・請求（税込）       : 目標値と完全一致で OK（許容差 0 円）
  ・支払（差引支給額） : 目標値との差が許容差（rates.yaml の pay_tolerance_yen）以内で OK
                        （源泉所得税は税額表の階級丸め等で数十円ずれ得るため）
  ・支払が NG のときは、本給／社保／社保控除後の内訳を出して原因を切り分ける
"""
import sys
import math
import argparse
import datetime as dt

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です:  pip install openpyxl --break-system-packages")
try:
    import yaml
except ImportError:
    sys.exit("PyYAML が必要です:  pip install pyyaml --break-system-packages")

# ---------------------------------------------------------------------------
# ▼ ここがYAML読み込み（旧: RATES辞書 と gensen_kou_2026関数 の直書きを置き換え）
#    料率・税額表・設定は rates.yaml にある。下の関数が読み込んで組み立てる。
# ---------------------------------------------------------------------------
RATES = {}          # load_config() で埋める
SETTINGS = {}       # 同上
GENSEN = {}         # 年度 -> 源泉計算関数

def load_config(path):
    """rates.yaml を読み、RATES / SETTINGS / GENSEN（モジュール全体で使う）を構築する。"""
    global RATES, SETTINGS, GENSEN
    with open(path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    ins = cfg["insurance"]
    yr = ins[ins["active_year"]]             # insurance.active_year で年度サブブロックを切替
    RATES = {
        "health": yr["health"],              # 健康保険（全47都道府県の辞書）
        "default_pref": ins.get("default_prefecture"),  # 都道府県不明時のフォールバック
        "kaigo": yr["care"],                 # 介護保険
        "kosei": yr["pension"],              # 厚生年金
        "koyo":  yr["employment"],           # 雇用保険（労働者負担）
        "shohi": yr["consumption"],          # 消費税
    }
    s = cfg["settings"]
    SETTINGS = {
        "round_unit":  s["round_unit_min"],
        "pay_tol":     s["pay_tolerance_yen"],
        "legal_daily": s["legal_daily_hours"],
    }
    GENSEN = {int(y): _make_gensen(tbl) for y, tbl in cfg["withholding"].items()}
    return cfg

def _make_gensen(table):
    """rates.yaml の withholding[年] の表から、源泉所得税（月額表甲欄・特例）の計算関数を作る。"""
    def gensen(A, dependents):
        if A <= 0:
            return 0
        # 第1表 給与所得控除（[Aの上限, flat/rate, 率, 加算] / 1円未満切上）
        kyuyo = None
        for hi, kind, rate, add in table["salary_deduction"]:
            if hi is None or A <= hi:
                kyuyo = add if kind == "flat" else A * rate + add
                break
        kyuyo = math.ceil(kyuyo)
        # 第2表 配偶者・扶養控除（1人あたり一定額）
        fuyo_ded = table["dependent_deduction"] * (dependents or 0)
        # 第3表 基礎控除（[Aの上限, 額]）
        kiso = 0
        for hi, amt in table["basic_deduction"]:
            if hi is None or A <= hi:
                kiso = amt
                break
        B = A - kyuyo - fuyo_ded - kiso       # 課税給与所得金額
        if B <= 0:
            return 0
        # 第4表 税額（[Bの上限, 税率, 控除額] / 10円未満四捨五入）
        for hi, rate, ded in table["tax_brackets"]:
            if hi is None or B <= hi:
                return int(round(B * rate - ded, -1))
        return 0
    return gensen

# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------
def headers(ws):
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

def rows_as_dicts(ws):
    H = headers(ws)
    out = []
    for r in range(2, ws.max_row + 1):
        out.append({H[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)})
    return out

def num(v):
    return v if isinstance(v, (int, float)) else 0

# ---------------------------------------------------------------------------
# 勤怠 → 時間区分／有給
# ---------------------------------------------------------------------------
def split_attendance(contract, recs, round_unit, legal_daily):
    shotei = contract["所定労働時間"]
    hours = {"base": 0.0, "nai": 0.0, "gai": 0.0}
    leave_pay = 0
    leave_bill = 0
    work_days = 0
    leave_days = 0
    trans_pay = 0
    for rec in recs:
        s, e = rec.get("開始日時"), rec.get("終了日時")
        is_leave = (rec.get("1日有給休暇フラグ") == 1) or (rec.get("半日有給休暇フラグ") == 1)
        if s is None or e is None:
            if is_leave or num(rec.get("支払い料金")) or num(rec.get("請求料金")):
                leave_pay += num(rec.get("支払い料金"))
                leave_bill += num(rec.get("請求料金"))
                leave_days += 1
            continue
        worked = (e - s).total_seconds() / 60 - num(rec.get("休憩時間（分）")) - num(rec.get("深夜休憩時間（分）"))
        worked = (worked // round_unit) * round_unit     # 切捨て丸め
        wh = worked / 60.0
        base = min(wh, shotei)
        over = max(0.0, wh - shotei)
        if shotei >= legal_daily:
            nai, gai = 0.0, over
        else:
            nai = min(over, legal_daily - shotei)
            gai = max(0.0, wh - legal_daily)
        hours["base"] += base
        hours["nai"] += nai
        hours["gai"] += gai
        work_days += 1
        trans_pay += num(rec.get("通勤交通費：支給"))
    return hours, leave_pay, leave_bill, work_days, leave_days, trans_pay

# ---------------------------------------------------------------------------
# 請求・支払の計算
# ---------------------------------------------------------------------------
def calc_contract(contract, recs, round_unit, legal_daily, tax_year):
    hours, leave_pay, leave_bill, wdays, ldays, trans_pay = \
        split_attendance(contract, recs, round_unit, legal_daily)

    # 請求(税抜) : 有給は対象外
    bill_pretax = (hours["base"] * num(contract["請求料金 基本単価"])
                   + hours["nai"] * num(contract["請求料金 法定内残業単価"])
                   + hours["gai"] * num(contract["請求料金 法定外残業単価"]))

    # 支払
    hongyu = (hours["base"] * num(contract["支払料金 基本単価"])
              + hours["nai"] * num(contract["支払料金 法定内残業単価"])
              + hours["gai"] * num(contract["支払料金 法定外残業単価"])
              + leave_pay)
    std = num(contract["標準報酬月額"])
    # 健康保険率は契約の「都道府県」で選ぶ（全47都道府県をrates.yamlに保持）
    pref = (contract.get("都道府県") or RATES["default_pref"])
    kenpo_rate = RATES["health"].get(pref)
    if kenpo_rate is None:                    # 末尾の都/道/府/県表記ゆれを吸収して再探索
        kenpo_rate = next((v for k, v in RATES["health"].items()
                           if pref and (k.rstrip("都道府県") == str(pref).rstrip("都道府県"))),
                          RATES["health"][RATES["default_pref"]])
    kenpo = round(std * (kenpo_rate + (RATES["kaigo"] if contract.get("介護保険加入") == 1 else 0)) / 2) \
        if contract.get("健康保険加入") == 1 else 0
    kosei = round(std * RATES["kosei"] / 2) if contract.get("厚生年金保険加入") == 1 else 0
    koyo = round((hongyu + trans_pay) * RATES["koyo"]) if contract.get("雇用保険加入") == 1 else 0
    shaho = kenpo + kosei + koyo
    after = hongyu - shaho
    gensen = GENSEN[tax_year](after, num(contract.get("扶養人数")))
    jumin = 0
    shikyu = after + trans_pay - gensen - jumin

    return {
        "name": contract.get("料金名"),
        "hours": hours, "leave_pay": leave_pay, "work_days": wdays, "leave_days": ldays,
        "bill_pretax": bill_pretax,
        "hongyu": hongyu, "trans": trans_pay,
        "kenpo": kenpo, "kosei": kosei, "koyo": koyo, "shaho": shaho,
        "after": after, "gensen": gensen, "shikyu": shikyu,
        "flags": (contract.get("健康保険加入"), contract.get("厚生年金保険加入"),
                  contract.get("介護保険加入"), contract.get("雇用保険加入")),
    }

# ---------------------------------------------------------------------------
# シート探索
# ---------------------------------------------------------------------------
def find_sheets(wb):
    contract_sheets, att_sheets = {}, {}
    for name in wb.sheetnames:
        mark = name[0] if name and name[0] in "①②③④⑤⑥⑦⑧⑨" else None
        if mark is None:
            continue
        if "契約" in name:
            contract_sheets[mark] = name
        elif "勤怠" in name:
            att_sheets[mark] = name
    cases = []
    for mark in sorted(set(contract_sheets) | set(att_sheets)):
        if mark in contract_sheets and mark in att_sheets:
            cases.append((mark, contract_sheets[mark], att_sheets[mark]))
    return cases

def parse_overview(wb):
    name = next((n for n in wb.sheetnames if "概要" in n), None)
    if not name:
        return {}
    ws = wb[name]
    targets, cur = {}, None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        if isinstance(a, str) and a and a[0] in "①②③④⑤⑥⑦⑧⑨" and isinstance(c, (int, float)):
            cur = a[0]
            targets[cur] = {"bill": c, "pay": []}
        if cur and isinstance(d, (int, float)):
            targets[cur]["pay"].append(d)
    return targets

# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel")
    ap.add_argument("--rates", default="rates.yaml", help="料率・税額表のYAMLファイル")
    ap.add_argument("--tax-year", type=int, default=2026)
    ap.add_argument("--round-unit", type=int, default=None, help="実働時間の切捨て単位（分）。未指定ならYAMLの値")
    ap.add_argument("--pay-tol", type=int, default=None, help="支払額の許容差（円）。未指定ならYAMLの値")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    load_config(args.rates)                              # ← ここでrates.yamlを読み込む
    round_unit = args.round_unit if args.round_unit is not None else SETTINGS["round_unit"]
    pay_tol = args.pay_tol if args.pay_tol is not None else SETTINGS["pay_tol"]
    legal_daily = SETTINGS["legal_daily"]

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    cases = find_sheets(wb)
    targets = parse_overview(wb)
    if not cases:
        sys.exit("契約/勤怠シート（①②…）が見つかりません。")

    all_ok = True
    print("=" * 78)
    print(f"検証対象: {args.excel}")
    print(f"設定: 料率={args.rates} / 税年度={args.tax_year} / 実働丸め={round_unit}分切捨 / 支払許容差±{pay_tol}円")
    print("=" * 78)

    for mark, cs, as_ in cases:
        contracts = {row["契約番号"]: row for row in rows_as_dicts(wb[cs]) if row.get("契約番号") is not None}
        att = {}
        for row in rows_as_dicts(wb[as_]):
            cn = row.get("契約番号")
            if cn is not None:
                att.setdefault(cn, []).append(row)

        results = []
        for cn, c in contracts.items():
            recs = att.get(cn, [])
            if recs:
                results.append((cn, calc_contract(c, recs, round_unit, legal_daily, args.tax_year)))

        bill_pretax = sum(r["bill_pretax"] for _, r in results)
        bill_total = int(round(bill_pretax * (1 + RATES["shohi"])))
        tgt = targets.get(mark, {})

        print(f"\n■ ケース {mark}")
        if "bill" in tgt:
            ok = (bill_total == tgt["bill"])
            all_ok &= ok
            print(f"  [請求] 計算={bill_total:,}円  目標={tgt['bill']:,}円  → {'OK' if ok else 'NG'}")
        else:
            print(f"  [請求] 計算={bill_total:,}円  （概要に目標値なし）")

        comp = sorted((r["shikyu"], cn, r) for cn, r in results)
        tg_pay = sorted(tgt.get("pay", []))
        used = [False] * len(tg_pay)
        print(f"  [支払] スタッフ別 差引支給額:")
        for shikyu, cn, r in comp:
            best, bestdiff = None, None
            for i, t in enumerate(tg_pay):
                if used[i]:
                    continue
                diff = abs(t - shikyu)
                if bestdiff is None or diff < bestdiff:
                    best, bestdiff = i, diff
            if best is not None:
                used[best] = True
                tval = tg_pay[best]
                ok = bestdiff <= pay_tol
                all_ok &= ok
                note = "" if ok else f"  差{shikyu - tval:+,}円"
                print(f"    契約{cn} {str(r['name'])[:14]:<14} 計算={shikyu:>9,.0f}円  目標={tval:>9,}円 → {'OK' if ok else 'NG'}{note}")
            else:
                print(f"    契約{cn} {str(r['name'])[:14]:<14} 計算={shikyu:>9,.0f}円  （対応する目標値なし）")
            if args.verbose or (best is not None and bestdiff > pay_tol):
                print(f"        ├ 本給{r['hongyu']:,.0f}(うち有給{r['leave_pay']:,}) 交通{r['trans']:,}"
                      f" 健保{r['kenpo']:,} 厚年{r['kosei']:,} 雇用{r['koyo']:,}")
                print(f"        ├ 社保控除後{r['after']:,.0f} 源泉{r['gensen']:,} 社保加入(健/厚/介/雇)={r['flags']}")
        sheet_order = [r["shikyu"] for _, r in results]
        if tgt.get("pay") and sorted(sheet_order) == sorted(tgt["pay"]) and sheet_order != tgt["pay"]:
            print("    ※ 値の集合は一致するが、概要シートの並び順が計算順と異なる（表示順の入替の可能性）")

    print("\n" + "=" * 78)
    print(f"総合判定: {'すべて OK' if all_ok else 'NG あり（上記参照）'}")
    print("=" * 78)
    sys.exit(0 if all_ok else 1)

if __name__ == "__main__":
    main()
